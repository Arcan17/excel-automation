"""
Generates a formatted Excel report with charts from SalesSummary data.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime
from processor import SalesSummary


# --- Color palette ---
COLOR_HEADER_BG = "2E4057"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ACCENT = "048A81"
COLOR_ACCENT_LIGHT = "C8F4F1"
COLOR_ALT_ROW = "F2F7F7"
COLOR_TOTAL_BG = "054A91"
COLOR_TOTAL_FG = "FFFFFF"
COLOR_SUMMARY_BG = "048A81"
COLOR_SUMMARY_FG = "FFFFFF"

NUMBER_FORMAT_CLP = '#,##0" CLP"'
NUMBER_FORMAT_INT = '#,##0'


def _header_style(ws, cell_ref: str, value: str) -> None:
    cell = ws[cell_ref]
    cell.value = value
    cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
    )


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int,
                     header_color: str = COLOR_HEADER_BG,
                     number_cols: list = None,
                     currency_cols: list = None) -> int:
    """Write dataframe to worksheet, returns last row used."""
    number_cols = number_cols or []
    currency_cols = currency_cols or []

    # Headers
    for col_idx, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=1):
        bg = COLOR_ALT_ROW if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, value in enumerate(row):
            cell = ws.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value,
            )
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")

            col_name = df.columns[col_idx]
            if col_name in currency_cols:
                cell.number_format = NUMBER_FORMAT_CLP
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in number_cols:
                cell.number_format = NUMBER_FORMAT_INT
                cell.alignment = Alignment(horizontal="right", vertical="center")

    return start_row + len(df)


def generate_report(summary: SalesSummary, output_path: str) -> None:
    wb = openpyxl.Workbook()

    _build_summary_sheet(wb, summary)
    _build_monthly_sheet(wb, summary)
    _build_products_sheet(wb, summary)
    _build_sellers_sheet(wb, summary)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"Informe generado: {output_path}")


def _build_summary_sheet(wb: openpyxl.Workbook, summary: SalesSummary) -> None:
    ws = wb.create_sheet("Resumen Ejecutivo", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("B2:G2")
    title = ws["B2"]
    title.value = "INFORME DE VENTAS 2024"
    title.font = Font(bold=True, size=18, color=COLOR_HEADER_FG)
    title.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:G3")
    subtitle = ws["B3"]
    subtitle.value = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    subtitle.font = Font(italic=True, size=10, color="666666")
    subtitle.alignment = Alignment(horizontal="center")

    # KPI cards
    kpis = [
        ("B", "Ingresos Totales", f"{summary.total_revenue:,.0f} CLP"),
        ("D", "Unidades Vendidas", f"{summary.total_units:,}"),
        ("F", "Transacciones", f"{summary.total_transactions:,}"),
    ]

    for col, label, value in kpis:
        label_cell = ws[f"{col}5"]
        label_cell.value = label
        label_cell.font = Font(bold=True, color=COLOR_SUMMARY_FG, size=10)
        label_cell.fill = PatternFill("solid", fgColor=COLOR_SUMMARY_BG)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 24

        val_cell = ws[f"{col}6"]
        val_cell.value = value
        val_cell.font = Font(bold=True, size=14, color=COLOR_HEADER_BG)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = _thin_border()
        ws.row_dimensions[6].height = 30

        next_col = get_column_letter(openpyxl.utils.column_index_from_string(col) + 1)
        ws.merge_cells(f"{col}5:{next_col}5")
        ws.merge_cells(f"{col}6:{next_col}6")

    # Top products mini-table
    ws["B8"].value = "Top 5 Productos"
    ws["B8"].font = Font(bold=True, size=12, color=COLOR_HEADER_BG)
    ws.row_dimensions[8].height = 20

    top5 = summary.top_products[["Producto", "Ventas_Totales", "Unidades"]].copy()
    top5.columns = ["Producto", "Ventas (CLP)", "Unidades"]
    _write_dataframe(ws, top5, start_row=9, start_col=2,
                     currency_cols=["Ventas (CLP)"], number_cols=["Unidades"])

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 18


def _build_monthly_sheet(wb: openpyxl.Workbook, summary: SalesSummary) -> None:
    ws = wb.create_sheet("Ventas Mensuales")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value = "VENTAS MENSUALES 2024"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    df = summary.monthly_sales.copy()
    df.columns = ["Mes", "Ventas Totales", "Unidades", "Transacciones"]

    last_row = _write_dataframe(
        ws, df, start_row=3, start_col=1,
        currency_cols=["Ventas Totales"],
        number_cols=["Unidades", "Transacciones"]
    )

    # Total row
    total_row = last_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=COLOR_TOTAL_FG)
    ws.cell(row=total_row, column=2, value=summary.total_revenue).number_format = NUMBER_FORMAT_CLP
    ws.cell(row=total_row, column=3, value=summary.total_units).number_format = NUMBER_FORMAT_INT
    ws.cell(row=total_row, column=4, value=summary.total_transactions).number_format = NUMBER_FORMAT_INT
    for col in range(1, 5):
        c = ws.cell(row=total_row, column=col)
        c.fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
        c.font = Font(bold=True, color=COLOR_TOTAL_FG)
        c.border = _thin_border()

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Ventas Mensuales (CLP)"
    chart.y_axis.title = "CLP"
    chart.x_axis.title = "Mes"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(df) - 1)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(df) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"F3")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16


def _build_products_sheet(wb: openpyxl.Workbook, summary: SalesSummary) -> None:
    ws = wb.create_sheet("Top Productos")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"].value = "TOP 5 PRODUCTOS POR VENTAS"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    df = summary.top_products.copy()
    df.columns = ["Producto", "Ventas Totales", "Unidades", "Transacciones"]

    _write_dataframe(
        ws, df, start_row=3, start_col=1,
        header_color=COLOR_ACCENT,
        currency_cols=["Ventas Totales"],
        number_cols=["Unidades", "Transacciones"]
    )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16


def _build_sellers_sheet(wb: openpyxl.Workbook, summary: SalesSummary) -> None:
    ws = wb.create_sheet("Por Vendedor")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"].value = "RENDIMIENTO POR VENDEDOR"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_HEADER_FG)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    df = summary.sales_by_seller.copy()
    df.columns = ["Vendedor", "Ventas Totales", "Unidades", "Transacciones"]

    last_row = _write_dataframe(
        ws, df, start_row=3, start_col=1,
        currency_cols=["Ventas Totales"],
        number_cols=["Unidades", "Transacciones"]
    )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
